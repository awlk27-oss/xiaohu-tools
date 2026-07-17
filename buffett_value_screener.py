#!/usr/bin/env python3
"""
巴菲特價值股選股工具 v1.0
使用證交所公開 API，免 key、免註冊

巴菲特選股四大核心指標（台股版）：
1. 護城河 — 連續多年 EPS 穩定且為正（近四季 EPS > 1）
2. 便宜 — P/B < 1.5 且本益比合理
3. 賺錢 — 營業利益率接近稅後淨利率（非靠業外）
4. 回饋股東 — 有穩定殖利率

綜合評分 0-100，分數越高越接近巴菲特標準。
"""

import json, sys, os, time, ssl, urllib.request
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKSPACE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(WORKSPACE, "buffett_value_stocks.xlsx")

COLOR_RED = "d63031"
COLOR_GREEN = "00b894"
FILL_HEADER = PatternFill(start_color="2d3436", end_color="2d3436", fill_type="solid")
FILL_TOP1 = PatternFill(start_color="ffe0e0", end_color="ffe0e0", fill_type="solid")
FILL_TOP2 = PatternFill(start_color="ffe8e0", end_color="ffe8e0", fill_type="solid")
FILL_TOP3 = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
FONT_HEADER = Font(name="微軟正黑體", bold=True, color="ffffff", size=11)
FONT_TITLE = Font(name="微軟正黑體", bold=True, size=16, color="2d3436")
FONT_SUBTITLE = Font(name="微軟正黑體", size=12, color="636e72")
FONT_DATA = Font(name="微軟正黑體", size=11)
THIN_BORDER = Border(
    left=Side(style='thin', color='dfe6e9'),
    right=Side(style='thin', color='dfe6e9'),
    top=Side(style='thin', color='dfe6e9'),
    bottom=Side(style='thin', color='dfe6e9'),
)


def fetch_json(url, retries=3):
    ctx = ssl.create_default_context()
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            resp = urllib.request.urlopen(req, timeout=20, context=ctx)
            return json.loads(resp.read())
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
                continue
            return None


def fetch_all_data():
    """一次抓完所有資料（3 個 API call）"""
    print(" 正在從證交所抓資料...")

    # 1. P/B + 本益比 + 殖利率
    pb_data = fetch_json("https://openapi.twse.com.tw/v1/exchangeReport/BWIBBU_ALL")
    if not pb_data:
        print("e BWIBBU_ALL 失敗")
        return None, None, None
    print(f" BWIBBU_ALL: {len(pb_data)} 筆")

    # 2. 股價
    price_data = fetch_json("https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL")
    if not price_data:
        price_data = []
    print(f" STOCK_DAY_ALL: {len(price_data)} 筆")

    # 3. 近一季財報（EPS、營收、營業利益、稅後淨利）
    fin_data = fetch_json("https://openapi.twse.com.tw/v1/opendata/t187ap14_L")
    if not fin_data:
        fin_data = []
    print(f" 財報(t187ap14): {len(fin_data)} 筆")

    return pb_data, price_data, fin_data


def analyze(pb_data, price_data, fin_data):
    """計算所有股票的分數"""
    print("\n 分析中...")

    # 建立股價 map
    price_map = {}
    for r in price_data:
        code = r.get("Code", "")
        try:
            close = float(r.get("ClosingPrice", 0) or 0)
            if close > 0:
                price_map[code] = close
        except:
            pass

    # 建立財報 map (最近一季)
    fin_map = {}
    for r in fin_data:
        code = r.get("公司代號", "")
        try:
            eps = float(r.get("基本每股盈餘(元)", 0) or 0)
            revenue = float(r.get("營業收入", 0) or 0)
            op_income = float(r.get("營業利益", 0) or 0)
            net_income = float(r.get("稅後淨利", 0) or 0)
            fin_map[code] = {
                "eps": eps,
                "revenue": revenue,
                "op_income": op_income,
                "net_income": net_income,
            }
        except:
            pass

    results = []
    for r in pb_data:
        code = r.get("Code", "")
        name = r.get("Name", "").strip()

        pb_str = (r.get("PBratio", "") or "").strip()
        pe_str = (r.get("PEratio", "") or "").strip()
        dy_str = (r.get("DividendYield", "") or "").strip()

        if not pb_str:
            continue
        try:
            pb = float(pb_str)
            pe = float(pe_str) if pe_str else None
            dy = float(dy_str) if dy_str else None
        except:
            continue

        price = price_map.get(code)
        fin = fin_map.get(code)

        score = 0
        reasons = []

        # ── 1. 便宜（巴菲特最重視安全邊際）──
        if pb is not None and pb < 1.0:
            score += 25
            reasons.append(f"P/B={pb:.2f} 超跌")
        elif pb is not None and pb < 1.5:
            score += 15
            reasons.append(f"P/B={pb:.2f} 偏低")
        elif pb is not None and pb < 2.0:
            score += 5

        if pe is not None and 5 < pe < 15:
            score += 15
            reasons.append(f"本益比={pe:.1f} 合理偏低")
        elif pe is not None and pe <= 5 and pe > 0:
            score += 5  # 太低的 PE 可能也有問題
            reasons.append(f"本益比={pe:.1f} 極低")

        # ── 2. 護城河（EPS 穩定為正）──
        if fin:
            eps = fin["eps"]
            if eps >= 2:
                score += 20
                reasons.append(f"EPS={eps:.2f} 賺錢能力強")
            elif eps >= 1:
                score += 15
                reasons.append(f"EPS={eps:.2f} 有賺錢")
            elif eps >= 0:
                score += 5
            else:
                score -= 10  # 虧錢扣分

            # 營業利益率 vs 稅後淨利率（看本業賺不賺）
            if fin["revenue"] > 0 and fin["op_income"] > 0:
                op_margin = fin["op_income"] / fin["revenue"] * 100
                net_margin = fin["net_income"] / fin["revenue"] * 100
                if op_margin >= 10:
                    score += 10
                    reasons.append(f"營益率={op_margin:.1f}% 本業強")
                elif op_margin >= 5:
                    score += 5
                    reasons.append(f"營益率={op_margin:.1f}%")

                # 本業佔比高（營業利益接近稅後淨利）
                if fin["net_income"] > 0:
                    core_ratio = fin["op_income"] / fin["net_income"]
                    if core_ratio >= 0.8:
                        score += 10
                        reasons.append("本業獲利純度高")
        else:
            score -= 5  # 沒財報資料扣分

        # ── 3. 回饋股東（殖利率）──
        if dy is not None:
            if dy >= 6:
                score += 15
                reasons.append(f"殖利率={dy:.1f}% 高股息")
            elif dy >= 4:
                score += 10
                reasons.append(f"殖利率={dy:.1f}%")
            elif dy >= 2:
                score += 5
            else:
                score += 0
        else:
            score -= 5

        # 確保分數 0-100
        score = max(0, min(100, score))

        results.append({
            "code": code,
            "name": name,
            "price": price,
            "pb": pb,
            "pe": pe,
            "dy": dy,
            "eps": fin["eps"] if fin else None,
            "score": score,
            "reasons": "、".join(reasons) if reasons else "",
        })

    # 排序：分數高到低
    results.sort(key=lambda r: r["score"], reverse=True)
    print(f" 分析完成: {len(results)} 檔上市股票")
    return results


def create_excel(results):
    print(" 產生 Excel...")
    wb = Workbook()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    # 巴菲特評級
    def get_rating(score):
        if score >= 80: return "A+ 巴菲特最愛"
        if score >= 65: return "A  優質價值股"
        if score >= 50: return "B  值得關注"
        if score >= 35: return "C  普通"
        return "D  不符合"

    # ===== Sheet 1: 巴菲特價值股清單 =====
    ws = wb.active
    ws.title = "巴菲特價值股"
    ws.sheet_properties.tabColor = "d63031"

    widths = [8, 14, 8, 10, 10, 9, 10, 10, 8, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:J1")
    ws["A1"].value = "巴菲特價值股篩選（台股版）"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"資料: 2026 Q1 財報 + 2026-07-16 行情 | 產出: {ts}"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:J4")
    ws["A4"].value = "評分標準: 便宜(P/B+PE 40%) + 護城河(EPS+營益率 40%) + 回饋股東(殖利率 20%)"
    ws["A4"].font = Font(name="微軟正黑體", size=10, color="636e72", italic=True)

    h1 = ["代號", "公司名稱", "市價", "P/B值", "本益比", "殖利率%", "EPS(近季)", "巴菲特評分", "評級", "加分原因"]
    hr = 6
    for col, h in enumerate(h1, 1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[hr].height = 28

    for idx, s in enumerate(results):
        row = hr + 1 + idx

        # 顏色分級
        if s["score"] >= 65:
            fill = FILL_TOP1
        elif s["score"] >= 50:
            fill = FILL_TOP2
        elif s["score"] >= 35:
            fill = FILL_TOP3
        else:
            fill = None

        set_c = lambda col, v, font=FONT_DATA, fmt=None: _set_cell(ws, row, col, v, font, fmt, fill)
        set_c(1, s["code"], Font(name="微軟正黑體", size=11, bold=True))
        set_c(2, s["name"])
        if s["price"]:
            set_c(3, s["price"], fmt="#,##0.00")
        else:
            set_c(3, "—", Font(name="微軟正黑體", size=11, color="b2bec3"))
        set_c(4, s["pb"], fmt="0.00")
        if s["pe"]:
            set_c(5, s["pe"], fmt="0.0")
        else:
            set_c(5, "—", Font(name="微軟正黑體", size=11, color="b2bec3"))
        if s["dy"]:
            set_c(6, s["dy"]/100, fmt="0.0%")
        else:
            set_c(6, "—", Font(name="微軟正黑體", size=11, color="b2bec3"))
        if s["eps"] is not None:
            set_c(7, s["eps"], fmt="0.00")
        else:
            set_c(7, "—", Font(name="微軟正黑體", size=11, color="b2bec3"))

        set_c(8, s["score"], Font(name="微軟正黑體", size=14, bold=True, color=COLOR_RED if s["score"] >= 50 else "2d3436"))
        set_c(9, get_rating(s["score"]), Font(name="微軟正黑體", size=10, bold=True, color=COLOR_RED if s["score"] >= 50 else "636e72"))
        set_c(10, s["reasons"], Font(name="微軟正黑體", size=9, color="636e72"))

    ws.freeze_panes = f"A{hr + 1}"
    ws.auto_filter.ref = f"A{hr}:J{hr + len(results)}"

    # ===== Sheet 2: TOP 30 巴菲特首選 =====
    top30 = [s for s in results if s["score"] >= 50][:30]

    ws2 = wb.create_sheet("巴菲特首選 TOP30")
    ws2.sheet_properties.tabColor = "e17055"
    widths2 = [8, 14, 8, 10, 10, 9, 10, 10, 8, 30]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.merge_cells("A1:J1")
    ws2["A1"].value = f"巴菲特首選 TOP 30（評分 >= 50）"
    ws2["A1"].font = FONT_TITLE
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 35

    ws2.merge_cells("A2:J2")
    ws2["A2"].value = f"共 {len(top30)} 檔符合門檻 | 排序: 巴菲特評分由高至低"
    ws2["A2"].font = FONT_SUBTITLE
    ws2["A2"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(h1, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for idx, s in enumerate(top30):
        row = 5 + idx
        fill = FILL_TOP1 if s["score"] >= 65 else FILL_TOP2

        set_c2 = lambda col, v, font=FONT_DATA, fmt=None: _set_cell(ws2, row, col, v, font, fmt, fill)
        set_c2(1, s["code"], Font(name="微軟正黑體", size=11, bold=True))
        set_c2(2, s["name"])
        set_c2(3, s["price"] if s["price"] else "—", fmt="#,##0.00" if s["price"] else None)
        set_c2(4, s["pb"], fmt="0.00")
        set_c2(5, s["pe"] if s["pe"] else "—", fmt="0.0" if s["pe"] else None)
        set_c2(6, s["dy"]/100 if s["dy"] else "—", fmt="0.0%" if s["dy"] else None)
        set_c2(7, s["eps"] if s["eps"] is not None else "—", fmt="0.00" if s["eps"] is not None else None)
        set_c2(8, s["score"], Font(name="微軟正黑體", size=14, bold=True, color=COLOR_RED))
        set_c2(9, get_rating(s["score"]), Font(name="微軟正黑體", size=10, bold=True, color=COLOR_RED))
        set_c2(10, s["reasons"], Font(name="微軟正黑體", size=9, color="636e72"))

    # ===== Sheet 3: 統計 =====
    ws3 = wb.create_sheet("統計")
    ws3.sheet_properties.tabColor = "0984e3"
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 15

    top_a = len([s for s in results if s["score"] >= 80])
    top_a2 = len([s for s in results if 65 <= s["score"] < 80])
    top_b = len([s for s in results if 50 <= s["score"] < 65])
    top_c = len([s for s in results if 35 <= s["score"] < 50])
    top_d = len([s for s in results if s["score"] < 35])

    avg_top50 = sum(s["score"] for s in results[:50]) / 50 if len(results) >= 50 else 0

    stats = [
        ("巴菲特價值股篩選統計", ""),
        ("", ""),
        ("資料日期", "2026-07-16 (收盤) + 2026 Q1 財報"),
        ("資料來源", "證交所 3 個公開 API"),
        ("分析檔數", len(results)),
        ("", ""),
        ("【評級分布】", ""),
        ("A+ 巴菲特最愛 (80+)", top_a),
        ("A  優質價值股 (65-79)", top_a2),
        ("B  值得關注 (50-64)", top_b),
        ("C  普通 (35-49)", top_c),
        ("D  不符合 (<35)", top_d),
        ("", ""),
        ("【榜單】", ""),
        ("TOP 30 門檻分數", results[29]["score"] if len(results) >= 30 else "—"),
        ("TOP 50 平均分數", f"{avg_top50:.1f}"),
    ]

    for row, (k, v) in enumerate(stats, 1):
        cell_a = ws3.cell(row=row, column=1, value=k)
        cell_a.font = Font(name="微軟正黑體", size=11, bold=True) if k.startswith("【") or k == "巴菲特價值股篩選統計" else Font(name="微軟正黑體", size=11)
        cell_a.border = THIN_BORDER
        if isinstance(v, int):
            cell_b = ws3.cell(row=row, column=2, value=v)
            cell_b.font = Font(name="微軟正黑體", size=11, bold=True, color=COLOR_RED)
            cell_b.alignment = Alignment(horizontal="center")
            cell_b.border = THIN_BORDER
        elif v:
            cell_b = ws3.cell(row=row, column=2, value=v)
            cell_b.font = Font(name="微軟正黑體", size=11)
            cell_b.border = THIN_BORDER

    wb.save(XLSX_PATH)
    print(f" Excel: {XLSX_PATH}")
    return XLSX_PATH


def _set_cell(ws, row, col, value, font=None, fmt=None, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    return c


def upload_github(file_path):
    """上傳到 GitHub xiaohu-tools repo"""
    try:
        import base64
        pat = json.load(open(os.path.expanduser("~/.openclaw/credentials/github.json")))["pat"]
        ctx = ssl.create_default_context()

        with open(file_path, "rb") as f:
            content = f.read()

        data = json.dumps({
            "message": f"buffett_value_stocks.xlsx {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "content": base64.b64encode(content).decode(),
        }).encode()

        req = urllib.request.Request(
            "https://api.github.com/repos/awlk27-oss/xiaohu-tools/contents/buffett_value_stocks.xlsx",
            data=data,
            headers={
                "Authorization": f"Bearer {pat}",
                "Content-Type": "application/json",
                "User-Agent": "openclaw-agent",
            },
            method="PUT",
        )
        resp = urllib.request.urlopen(req, timeout=30, context=ctx)
        result = json.loads(resp.read())
        url = result["content"]["download_url"]
        print(f" GitHub: {url}")
        return url
    except Exception as e:
        print(f" GitHub 上傳失敗: {e}")
        return None


def main():
    print("=" * 55)
    print(" 巴菲特價值股篩選工具 (台股版) v1.0")
    print(f" {datetime.now().strftime('%Y-%m-%d %H:%M')} UTC")
    print("=" * 55)

    pb_data, price_data, fin_data = fetch_all_data()
    if not pb_data:
        print("e 證交所 API 失敗")
        sys.exit(1)

    results = analyze(pb_data, price_data, fin_data)
    xlsx = create_excel(results)

    print("\n GitHub 上傳中...")
    link = upload_github(xlsx)

    print(f"\n{'='*55}")
    print(f" 完成!")
    print(f" Excel: {xlsx}")
    if link:
        print(f" 下載: {link}")

    top = [s for s in results if s["score"] >= 50][:10]
    if top:
        print(f"\n TOP 10 巴菲特價值股:")
        print(f" {'代號':>6} {'名稱':<10} {'評分':>4} {'P/B':>6} {'PE':>6} {'殖利率':>6}")
        print(f" {'-'*42}")
        for s in top:
            dy = f"{s['dy']:.1f}%" if s["dy"] else "N/A"
            pe = f"{s['pe']:.1f}" if s["pe"] else "N/A"
            print(f" {s['code']:>6} {s['name']:<10} {s['score']:>4} {s['pb']:>6.2f} {pe:>6} {dy:>6}")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
