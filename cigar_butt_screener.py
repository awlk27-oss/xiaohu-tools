#!/usr/bin/env python3
"""
台股菸屁股選股工具 v1.1
使用證交所公開 API，免 API key、免註冊

功能：
1. 從證交所抓上市股票 P/B 資料 + 股價
2. 篩選 P/B < 1（菸屁股候選）
3. 輸出 Excel 表格（3 個 Sheet），自動上傳 Dropbox

用法：
  python3 cigar_butt_screener.py
"""

import json, sys, os, time, ssl, urllib.request, urllib.parse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKSPACE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(WORKSPACE, "cigar_butt_results.xlsx")

COLOR_RED = "d63031"
FILL_RED = PatternFill(start_color="ffe0e0", end_color="ffe0e0", fill_type="solid")
FILL_HEADER = PatternFill(start_color="2d3436", end_color="2d3436", fill_type="solid")
FONT_HEADER = Font(name="微軟正黑體", bold=True, color="ffffff", size=12)
FONT_TITLE = Font(name="微軟正黑體", bold=True, size=16, color="2d3436")
FONT_SUBTITLE = Font(name="微軟正黑體", size=12, color="636e72")
FONT_DATA = Font(name="微軟正黑體", size=11)
FONT_SECTION = Font(name="微軟正黑體", bold=True, size=13, color="d63031")
THIN_BORDER = Border(
    left=Side(style='thin', color='dfe6e9'),
    right=Side(style='thin', color='dfe6e9'),
    top=Side(style='thin', color='dfe6e9'),
    bottom=Side(style='thin', color='dfe6e9'),
)


def fetch_json(url, retries=3):
    ctx = ssl.create_default_context()
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            resp = urllib.request.urlopen(req, timeout=20, context=ctx)
            return json.loads(resp.read())
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
                continue
            print(f"  e {e}")
            return None


def fetch_twse_pb():
    """抓證交所 P/B + 殖利率 + 本益比"""
    print(" 正在抓取證交所資料...")
    data = fetch_json("https://openapi.twse.com.tw/v1/exchangeReport/BWIBBU_ALL")
    if not data:
        print("e 證交所 API 失敗")
        return [], []

    stocks = []
    for r in data:
        pb_str = (r.get("PBratio", "") or "").strip()
        if not pb_str:
            continue
        try:
            pb = float(pb_str)
        except:
            continue
        dy = (r.get("DividendYield", "") or "").strip()
        pe = (r.get("PEratio", "") or "").strip()
        stocks.append({
            "code": r.get("Code", ""),
            "name": r.get("Name", "").strip(),
            "pb": pb,
            "dy": float(dy) if dy else None,
            "pe": float(pe) if pe else None,
        })

    cigar = [s for s in stocks if s["pb"] < 1.0]
    cigar.sort(key=lambda r: r["pb"])
    print(f" 上市共 {len(stocks)} 檔有 P/B 資料")
    print(f" P/B < 1: {len(cigar)} 檔")
    return cigar, stocks


def fetch_prices(cigar):
    """從證交所抓個股收盤價"""
    print(" 抓取股價...")
    url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
    data = fetch_json(url)
    price_map = {}
    if data:
        for r in data:
            code = r.get("Code", "")
            try:
                close = float(r.get("ClosingPrice", 0) or 0)
                if close > 0:
                    price_map[code] = close
            except:
                pass
    for s in cigar:
        s["price"] = price_map.get(s["code"])
    have = sum(1 for s in cigar if s["price"])
    print(f" {have}/{len(cigar)} 檔有股價")
    return cigar


def set_cell(ws, row, col, value, font=None, fmt=None, fill=None, align_center=True):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if align_center: c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    return c


def create_excel(cigar, all_stocks):
    print(" 產生 Excel...")
    wb = Workbook()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    decent = [s for s in cigar
              if 0.5 <= s["pb"] <= 0.85
              and s["dy"] is not None and s["dy"] >= 4]
    decent.sort(key=lambda r: r["dy"], reverse=True)

    # ===== Sheet 1: 清單 =====
    ws = wb.active
    ws.title = "清單"
    ws.sheet_properties.tabColor = "d63031"
    widths1 = [8, 14, 8, 10, 11, 11, 18, 20]
    for i, w in enumerate(widths1, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:H1")
    ws["A1"].value = "台股菸屁股篩選 (P/B < 1)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"資料日期: 2026-07-16 (上市公司) | 產出: {ts}"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:H4")
    ws["A4"].value = "條件: P/B < 1.0 (股價低於每股淨值) | 來源: 證交所 BWIBBU_ALL"
    ws["A4"].font = Font(name="微軟正黑體", size=10, color="636e72", italic=True)

    h1 = ["代號", "股票名稱", "市場", "P/B值", "殖利率%", "本益比", "潛力評價", "備註"]
    hr = 6
    for col, h in enumerate(h1, 1):
        set_cell(ws, hr, col, h, font=FONT_HEADER, fill=FILL_HEADER)
    ws.row_dimensions[hr].height = 28

    for idx, s in enumerate(cigar):
        row = hr + 1 + idx
        set_cell(ws, row, 1, s["code"], font=Font(name="微軟正黑體", size=11, bold=True))
        set_cell(ws, row, 2, s["name"], font=FONT_DATA)
        set_cell(ws, row, 3, "上市", font=FONT_DATA)

        pb_font = Font(name="微軟正黑體", size=12, bold=True, color=COLOR_RED if s["pb"] < 0.5 else "2d3436")
        set_cell(ws, row, 4, s["pb"], font=pb_font, fmt="0.00")

        if s["dy"] is not None:
            dy_font = Font(name="微軟正黑體", size=11, color=COLOR_RED if s["dy"] >= 6 else "2d3436")
            set_cell(ws, row, 5, s["dy"]/100, font=dy_font, fmt="0.00%")
        else:
            set_cell(ws, row, 5, "—", font=Font(name="微軟正黑體", size=11, color="b2bec3"))

        if s["pe"] is not None:
            set_cell(ws, row, 6, s["pe"], font=FONT_DATA, fmt="0.0")
        else:
            set_cell(ws, row, 6, "—", font=Font(name="微軟正黑體", size=11, color="b2bec3"))

        has_dy = s["dy"] is not None and s["dy"] >= 4
        if s["pb"] < 0.4 and not has_dy:
            set_cell(ws, row, 7, "價值陷阱風險高", font=Font(name="微軟正黑體", size=10, color="d63031"))
        elif s["pb"] <= 0.85 and has_dy:
            set_cell(ws, row, 7, "有料", font=Font(name="微軟正黑體", size=10, bold=True, color="d63031"))
            ws.cell(row=row, column=1).fill = FILL_RED
        elif s["pb"] <= 0.65:
            set_cell(ws, row, 7, "可觀察", font=Font(name="微軟正黑體", size=10, color="636e72"))
        else:
            set_cell(ws, row, 7, "", font=Font(name="微軟正黑體", size=10, color="b2bec3"))

        notes = []
        if s["dy"] is not None and s["dy"] >= 8: notes.append(f"高殖利率{s['dy']:.1f}%")
        if s["pe"] is not None and 0 < s["pe"] < 10: notes.append(f"低本益比{s['pe']:.1f}")
        if s["pb"] < 0.5: notes.append("嚴重破淨")
        set_cell(ws, row, 8, "、".join(notes) if notes else "", font=Font(name="微軟正黑體", size=10, color="636e72"))

    ws.freeze_panes = f"A{hr + 1}"
    ws.auto_filter.ref = f"A{hr}:H{hr + len(cigar)}"

    # ===== Sheet 2: 有料菸屁股 =====
    ws2 = wb.create_sheet("有料菸屁股")
    ws2.sheet_properties.tabColor = "e17055"
    widths2 = [8, 14, 8, 11, 10, 11, 11, 20]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.merge_cells("A1:H1")
    ws2["A1"].value = "有料的菸屁股"
    ws2["A1"].font = FONT_TITLE
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 35

    ws2.merge_cells("A2:H2")
    ws2["A2"].value = f"P/B 0.5~0.85 + 殖利率 >= 4% | 共 {len(decent)} 檔"
    ws2["A2"].font = FONT_SUBTITLE
    ws2["A2"].alignment = Alignment(horizontal="center")

    h2 = ["代號", "股票名稱", "市場", "收盤價", "P/B值", "殖利率", "本益比", "備註"]
    hr2 = 4
    for col, h in enumerate(h2, 1):
        set_cell(ws2, hr2, col, h, font=FONT_HEADER, fill=FILL_HEADER)

    for idx, s in enumerate(decent):
        row = hr2 + 1 + idx
        set_cell(ws2, row, 1, s["code"], font=Font(name="微軟正黑體", size=11, bold=True))
        set_cell(ws2, row, 2, s["name"], font=FONT_DATA)
        set_cell(ws2, row, 3, "上市", font=FONT_DATA)

        if s.get("price"):
            set_cell(ws2, row, 4, s["price"], font=Font(name="微軟正黑體", size=11, bold=True), fmt="#,##0.00")
        else:
            set_cell(ws2, row, 4, "—", font=Font(name="微軟正黑體", size=11, color="b2bec3"))

        set_cell(ws2, row, 5, s["pb"], font=Font(name="微軟正黑體", size=12, bold=True, color=COLOR_RED), fmt="0.00")
        set_cell(ws2, row, 6, s["dy"]/100, font=Font(name="微軟正黑體", size=11, bold=True, color=COLOR_RED), fmt="0.0%")

        if s["pe"] is not None:
            set_cell(ws2, row, 7, s["pe"], font=FONT_DATA, fmt="0.0")
        else:
            set_cell(ws2, row, 7, "—", font=Font(name="微軟正黑體", size=11, color="b2bec3"))

        pe = s["pe"] or 0
        n2 = []
        if s["dy"] >= 8: n2.append(f"高殖利率{s['dy']:.1f}%")
        if 0 < pe < 10: n2.append("低本益比")
        set_cell(ws2, row, 8, "、".join(n2) if n2 else "",
                 font=Font(name="微軟正黑體", size=10, color="636e72"))

    # ===== Sheet 3: 統計 =====
    ws3 = wb.create_sheet("統計")
    ws3.sheet_properties.tabColor = "0984e3"
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 15

    stats = [
        ("統計摘要", ""),
        ("", ""),
        ("資料日期", "2026-07-16"),
        ("資料來源", "證交所 BWIBBU_ALL + STOCK_DAY_ALL"),
        ("上市有 P/B 資料", len(all_stocks)),
        ("P/B < 1 (菸屁股)", len(cigar)),
        ("有料菸屁股 (P/B 0.5~0.85 + 殖利率>=4%)", len(decent)),
        ("", ""),
        ("P/B 分布", ""),
        ("P/B < 0.3", len([s for s in cigar if s["pb"] < 0.3])),
        ("0.3 <= P/B < 0.5", len([s for s in cigar if 0.3 <= s["pb"] < 0.5])),
        ("0.5 <= P/B < 0.7", len([s for s in cigar if 0.5 <= s["pb"] < 0.7])),
        ("0.7 <= P/B < 0.9", len([s for s in cigar if 0.7 <= s["pb"] < 0.9])),
        ("0.9 <= P/B < 1.0", len([s for s in cigar if 0.9 <= s["pb"] < 1.0])),
    ]

    for row, (k, v) in enumerate(stats, 1):
        cell_a = set_cell(ws3, row, 1, k, font=FONT_SECTION if k.endswith("分布") or k == "統計摘要" else Font(name="微軟正黑體", size=11))
        if k == "":
            continue
        if isinstance(v, int):
            set_cell(ws3, row, 2, v, font=Font(name="微軟正黑體", size=11, bold=True, color=COLOR_RED))

    wb.save(XLSX_PATH)
    print(f" Excel: {XLSX_PATH}")
    return XLSX_PATH


def upload_dropbox(file_path):
    creds_path = os.path.expanduser("~/.openclaw/credentials/dropbox.json")
    if not os.path.exists(creds_path):
        print(" 無 Dropbox 憑證，跳過上傳")
        return None

    with open(creds_path) as f:
        creds = json.load(f)
    ctx = ssl.create_default_context()
    filename = os.path.basename(file_path)
    dropbox_path = f"/小虎歌曲/台股分析/{filename}"

    def do_upload(token):
        with open(file_path, "rb") as f:
            data = f.read()
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream",
            "Dropbox-API-Arg": json.dumps({"path": dropbox_path, "mode": "overwrite"}),
        }
        req = urllib.request.Request("https://content.dropboxapi.com/2/files/upload", data=data, headers=headers, method="POST")
        return urllib.request.urlopen(req, timeout=30, context=ctx)

    try:
        do_upload(creds["access_token"])
        print(f" Dropbox: {dropbox_path}")
    except urllib.error.HTTPError as e:
        if e.code == 401 and creds.get("refresh_token"):
            print(" Token 過期，自動 refresh...")
            data = urllib.parse.urlencode({
                "grant_type": "refresh_token", "refresh_token": creds["refresh_token"],
                "client_id": "tj7phfpix5220nn", "client_secret": "***",
            }).encode()
            req = urllib.request.Request("https://api.dropboxapi.com/oauth2/token", data=data, method="POST")
            resp = urllib.request.urlopen(req, timeout=15, context=ctx)
            result = json.loads(resp.read())
            creds["access_token"] = result["access_token"]
            if result.get("refresh_token"): creds["refresh_token"] = result["refresh_token"]
            with open(creds_path, "w") as f: json.dump(creds, f, indent=2)
            do_upload(creds["access_token"])
            print(f" Dropbox (refresh): {dropbox_path}")
        else:
            print(f" e Dropbox 失敗: HTTP {e.code}")
            return None

    try:
        headers = {"Authorization": f"Bearer {creds['access_token']}", "Content-Type": "application/json"}
        req = urllib.request.Request(
            "https://api.dropboxapi.com/2/sharing/create_shared_link_with_settings",
            data=json.dumps({"path": dropbox_path, "settings": {"requested_visibility": "public"}}).encode(),
            headers=headers, method="POST",
        )
        resp = urllib.request.urlopen(req, timeout=15, context=ctx)
        link = json.loads(resp.read())["url"].replace("dl=0", "dl=0")
        print(f" {link}")
        return link
    except urllib.error.HTTPError as e:
        if e.code == 409:
            req2 = urllib.request.Request(
                "https://api.dropboxapi.com/2/sharing/list_shared_links",
                data=json.dumps({"path": dropbox_path}).encode(), headers=headers, method="POST",
            )
            links = json.loads(urllib.request.urlopen(req2, timeout=15, context=ctx).read()).get("links", [])
            if links:
                link = links[0]["url"].replace("dl=0", "dl=0")
                print(f" {link}")
                return link
        print(f" e 分享連結失敗: HTTP {e.code}")
        return None


def main():
    print("=" * 50)
    print(" 台股菸屁股選股工具 v1.1")
    print(f" {datetime.now().strftime('%Y-%m-%d %H:%M')} UTC")
    print("=" * 50)

    cigar, all_stocks = fetch_twse_pb()
    if not all_stocks:
        print("e 無法獲取資料")
        sys.exit(1)

    cigar = fetch_prices(cigar)
    create_excel(cigar, all_stocks)

    print("\n☁️ 上傳 Dropbox...")
    link = upload_dropbox(XLSX_PATH)

    decent = [s for s in cigar if 0.5 <= s["pb"] <= 0.85 and s["dy"] is not None and s["dy"] >= 4]
    decent.sort(key=lambda r: r["dy"], reverse=True)

    print(f"\n{'='*50}")
    print(f" 完成!")
    print(f" P/B < 1 菸屁股: {len(cigar)} 檔")
    print(f" 有料菸屁股: {len(decent)} 檔")
    print(f" Excel: {XLSX_PATH}")
    if link: print(f" Dropbox: {link}")
    print(f"{'='*50}")

    if decent:
        print(f"\nTOP 5 有料菸屁股:")
        for s in decent[:5]:
            ps = f" 市價{s['price']:.2f}" if s.get("price") else ""
            print(f"  {s['code']} {s['name']} P/B={s['pb']:.2f} 殖利率={s['dy']:.1f}% {ps}")


if __name__ == "__main__":
    main()
